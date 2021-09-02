from openpyxl import Workbook, load_workbook


def excel__init__():
    global wb, ws
    wb = Workbook()
    ws = wb.active
    ws.title = "Egzerisz Verileri"


def create_excel(regester_data, exercise_list):
    excel__init__()
    ws["A1"] = "Ad :"
    ws["B1"] = regester_data[2]
    ws["A2"] = "Soyad :"
    ws["B2"] = regester_data[3]
    ws["E1"] = "Tc :"
    ws["F1"] = regester_data[4]
    ws["E2"] = "Yaş :"
    ws["F2"] = regester_data[5]
    ws["H1"] = "Boy :"
    ws["I1"] = regester_data[6]
    ws["H2"] = "Kilo :"
    ws["I2"] = regester_data[7]

    ws["A4"] = "EGZERSİZ TARİHİ"
    ws["B4"] = "B EGZERSİZİ DURUM"
    ws["C4"] = "B EGZERSİZİ DURUM"
    ws["D4"] = "C EGZERSİZİ DURUM"

    for i in range(len(exercise_list)):
        ws.append([exercise_list[i][1], exercise_list[i][2], exercise_list[i][3], exercise_list[i][4]])

    wb.save("Egzersiz_Verileri.xlsx")
